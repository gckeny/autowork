import csv
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime, time


def r_exl(sheet_name, date1=''):
    inwb = load_workbook('22.xlsx')
    sheet = inwb[sheet_name]
    cols = sheet['b']  # 统计B列录入日期的障碍数量，以做编码用
    a = []
    for col in cols:
        a.append(col.value)
    return a.count(date1)


def rw_exl(sheet_name, zawt='', pzr='', date1='', zabm=''):
    inwb = load_workbook('22.xlsx')  # (r'C:\Users\Administrator\Desktop\抢修台账\20.xlsx')
    # print(inwb.sheetnames)
    sheet = inwb[sheet_name]  # 打开工作表
    # sheet=inwb.get_sheet_by_name('明德')
    # print(sheet['e4'].value)
    # x=0
    # for i in sheet['e']:
    #     if i.value==None:
    #         break
    #     else:
    #         x+=1
    # #print(x)
    # col_b='b'+str(x+1)
    # col_e='e'+str(x+1)
    # col_f='f'+str(x+1)
    # if date1=='':
    #     date1 = datetime.datetime.now().strftime("%Y-%m-%d")

    # sheet[col_e] = zawt
    # sheet[col_f] = pzr
    # time.sleep(1)
    # print(datetime.datetime.now().strftime("%Y-%m-%d"))

    # sheet.append(['', date1, '','',zawt,pzr])
    sheet.append({'a': zabm, 'b': date1, 'e': zawt, 'f': pzr})

    inwb.save("22.xlsx")


def auto_weixin():
    pass


while True:
    # try:

    print('障碍登记系统：')
    sheet_name = input('请输入抢修工程队(1明德,2传输局)：')
    if sheet_name not in ['1', '2']:  # != '1' and sheet_name != '2':
        print('输入错误，重新输入！')
        time.sleep(1)
        continue
    if sheet_name == '1':
        sheet_name = 'Sheet1'  # '明德'
    else:
        sheet_name = '传输局'

    date1 = input('请输入派障日期(不输入默认为当天)：')
    if '/' in date1:
        if date1.count('/') == 2:
            y, m, d = date1.split('/')
            zabm = y + m.zfill(2) + d.zfill(2)
            date1 = y + '-' + m.zfill(2) + '-' + d.zfill(2)
        elif date1.count('/') == 1:
            m, d = date1.split('/')
            zabm = datetime.datetime.now().strftime('%Y') + m.zfill(2) + d.zfill(2)
            date1 = datetime.datetime.now().strftime('%Y') + '-' + m.zfill(2) + '-' + d.zfill(2)
    elif '-' in date1:
        if date1.count('-') == 2:
            y, m, d = date1.split('-')
            zabm = y + m.zfill(2) + d.zfill(2)
            date1 = y + '-' + m.zfill(2) + '-' + d.zfill(2)
        elif date1.count('-') == 1:
            m, d = date1.split('-')
            zabm = datetime.datetime.now().strftime('%Y') + m.zfill(2) + d.zfill(2)
            date1 = datetime.datetime.now().strftime('%Y') + '-' + m.zfill(2) + '-' + d.zfill(2)
    elif date1 == '':
        date1 = datetime.datetime.now().strftime("%Y-%m-%d")
        zabm = datetime.datetime.now().strftime('%Y%m%d')
    else:
        print('输入错误，请重新输入！！！！')
        time.sleep(1)
        continue

    bm_hz = str(r_exl(sheet_name, date1=date1) + 1)
    zabm = zabm + bm_hz.zfill(2)

    zawt = input('请输入障碍问题：')
    pzr = input('请输入派障人：')

    auweixin = input('是否自动发微信通知工程队(默认不发送，Y/y发送)：')

    rw_exl(sheet_name, zawt=zawt, pzr=pzr, date1=date1, zabm=zabm)  # 调用读写EXCEL函数rw_exl()进行数据保存

    if auweixin == 'y' or auweixin == 'Y':
        auto_weixin()

    print('{0},{2}通知的需{3}维修的{1},本障碍已录入成功!'.format(zabm, zawt, pzr, sheet_name) )
    yn = input('是否继续输入(N/n退出，其他继续):')
    if yn == 'n' or yn == 'N':
        break
    # except:
    #     yn=input('******输入数据有误，是否继续输入!!!!!(N/n退出，其他继续):')
    #     if yn=='n' or yn=='N':
    #         break
